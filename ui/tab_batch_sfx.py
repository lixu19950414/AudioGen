"""Tab — 批量音效合成"""

from __future__ import annotations

import datetime
import tempfile
import zipfile
from pathlib import Path

import gradio as gr
import pandas as pd
from openpyxl.comments import Comment

from core.audio_utils import AudioFormat
from core.batch_processor import load_table
from core.app_logger import log_event, EVENT_BATCH
from ui.common import (
    BATCH_OUTPUT_DIR,
    batch_processor,
    task_queue,
)

_SFX_TEMPLATE_COLUMNS = ["prompt", "negative_prompt", "duration", "filename"]

_SFX_TEMPLATE_COMMENTS = {
    "prompt": "音效描述（必填）\n用英文描述想要的音效，\n如 gunshot, footsteps on gravel。",
    "negative_prompt": "负面提示词（可选）\n不希望出现的音效特征，\n如 music, voice。不填则不限制。",
    "duration": "时长秒数（可选）\n音效持续时间，1~47 秒。\n不填则使用页面上的默认时长。",
    "filename": "输出文件名（可选）\n不填则按行号自动命名，如 0001.wav。\n无需填写扩展名。",
}

_SFX_TEMPLATE_EXAMPLE = {
    "prompt": "gunshot in a large hall with echo",
    "negative_prompt": "music",
    "duration": 5,
    "filename": "gunshot_hall",
}


def _generate_sfx_template() -> str:
    """生成批量音效合成 Excel 模板文件，返回路径。"""
    df = pd.DataFrame([_SFX_TEMPLATE_EXAMPLE], columns=_SFX_TEMPLATE_COLUMNS)
    path = Path(tempfile.gettempdir()) / "批量音效合成模板.xlsx"
    df.to_excel(path, index=False)

    from openpyxl import load_workbook
    wb = load_workbook(path)
    ws = wb.active
    for col_idx, col_name in enumerate(_SFX_TEMPLATE_COLUMNS, start=1):
        if col_name in _SFX_TEMPLATE_COMMENTS:
            ws.cell(row=1, column=col_idx).comment = Comment(
                _SFX_TEMPLATE_COMMENTS[col_name], "AudioGen"
            )
    wb.save(path)
    return str(path)


def tab_batch_sfx():
    with gr.Tab("批量处理音效"):
        gr.Markdown(
            "### 上传 Excel 批量合成音效\n"
            "Excel 列：`prompt`（英文音效描述，必填）、`negative_prompt`（负面提示词，可选）、"
            "`duration`（时长秒数，可选）、`filename`（输出文件名，可选）"
        )
        with gr.Row():
            with gr.Column():
                template_btn = gr.Button("下载 Excel 模板")
                template_file = gr.File(label="模板文件", interactive=False)
                file_upload = gr.File(
                    label="上传 CSV / Excel 文件",
                    file_types=[".csv", ".xlsx", ".xls"],
                )
            with gr.Column():
                with gr.Row():
                    default_duration = gr.Slider(
                        minimum=1, maximum=47, value=10, step=1,
                        label="默认时长（秒）",
                    )
                    steps_slider = gr.Slider(
                        minimum=50, maximum=200, value=100, step=10,
                        label="推理步数",
                    )
                    guidance_slider = gr.Slider(
                        minimum=1, maximum=15, value=7, step=0.5,
                        label="引导系数",
                    )
                batch_fmt = gr.Radio(choices=["WAV", "MP3"], value="WAV", label="输出格式")
                batch_btn = gr.Button("开始批量合成", variant="primary")

        batch_log = gr.Textbox(label="处理日志", lines=15, interactive=False, max_lines=30)
        batch_download = gr.File(label="下载压缩包", interactive=False)

        def download_template():
            return _generate_sfx_template()

        def run_batch_sfx(file, fmt, duration, steps, guidance, request: gr.Request):
            user = (request.username or "unknown") if request else "-"
            if file is None:
                return "请先上传文件", None
            try:
                df = load_table(file.name)
            except Exception as e:
                return f"文件读取失败：{e}", None

            # 检查必要列
            if "prompt" not in df.columns:
                return "文件缺少必要列：prompt", None

            # 预检查
            errors: list[str] = []
            empty_rows: list[int] = []
            for idx, row in df.iterrows():
                row_num = int(idx) + 1
                prompt = str(row.get("prompt", "")).strip()
                if not prompt:
                    empty_rows.append(row_num)

            if empty_rows:
                errors.append(f"以下行 prompt 为空：第 {', '.join(map(str, empty_rows))} 行")

            if errors:
                return (
                    "文件预检查未通过，请修正后重试：\n\n" + "\n".join(errors),
                    None,
                )

            # 创建输出目录
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            file_stem = Path(file.orig_name).stem if hasattr(file, "orig_name") else Path(file.name).stem
            safe_stem = "".join(c if c.isalnum() or c in ("_", "-") else "_" for c in file_stem)
            folder_name = f"{timestamp}_{safe_stem}_sfx"
            out_dir = BATCH_OUTPUT_DIR / folder_name
            out_dir.mkdir(parents=True, exist_ok=True)

            log_event(EVENT_BATCH, f"开始音效批量 文件={file_stem} 总数={len(df)} 格式={fmt}", user=user)

            log_lines: list[str] = []

            def progress_cb(cur, total, msg):
                log_lines.append(msg)

            out_fmt: AudioFormat = "mp3" if fmt == "MP3" else "wav"
            task_desc = f"批量音效: {file_stem} ({len(df)} 条)"

            def do_batch():
                return batch_processor.run_sfx(
                    df=df,
                    output_dir=str(out_dir),
                    output_format=out_fmt,
                    default_duration=duration,
                    steps=int(steps),
                    guidance_scale=guidance,
                    progress_cb=progress_cb,
                )

            result = task_queue.submit(user, "batch", task_desc, do_batch)
            log_lines.append(result.summary())
            log_event(EVENT_BATCH, f"完成音效批量 文件={file_stem} 成功={result.success} 失败={result.failed} 总数={result.total}", user=user)

            # 压缩输出目录
            if result.success > 0:
                zip_path = BATCH_OUTPUT_DIR / f"{folder_name}.zip"
                with zipfile.ZipFile(str(zip_path), "w", zipfile.ZIP_DEFLATED) as zf:
                    for f in out_dir.iterdir():
                        if f.is_file():
                            zf.write(f, f.name)
                log_lines.append(f"\n压缩包已生成：{zip_path.name}")
                return "\n".join(log_lines), str(zip_path)

            return "\n".join(log_lines), None

        template_btn.click(fn=download_template, inputs=[], outputs=[template_file])
        batch_btn.click(
            fn=run_batch_sfx,
            inputs=[file_upload, batch_fmt, default_duration, steps_slider, guidance_slider],
            outputs=[batch_log, batch_download],
        )
